
from openpyxl import Workbook
from openpyxl import load_workbook
import os


def check_exists():
    return os.path.exists('tasks.xlsx')

def get_first_nonpopulated(sheet):
    column = sheet['A'] 
    for cell in column[::-1]:
        if cell.value is None:
            return cell.row
    return 1

def main_program():
    check_exist = check_exists()
    if check_exist:
        workbook = load_workbook('tasks.xlsx')
        sheet = workbook.active
        
        first_nonpopulated = get_first_nonpopulated(sheet) + 2
        
        task = input("Enter the task: ")
        date = input("Enter the date (YYYY-MM-DD): ")
        time = input("Enter the time (HH:MM): ")
        room = input("Enter the room number:")
        sheet.cell(row = first_nonpopulated, column = 1, value = task)
        sheet.cell(row = first_nonpopulated, column = 2, value = date)
        sheet.cell(row = first_nonpopulated, column = 3, value = time)
        sheet.cell(row = first_nonpopulated, column = 4, value = room)
        
        workbook.save('tasks.xlsx')
        
    else:
        workbook = Workbook()
        sheet = workbook.active
            
        
        sheet['A1'] = "Task"
        sheet['B1'] = "Date"
        sheet['C1'] = "Time"
        sheet['D1'] = "Room"
        
        task = input("Enter the task: ")
        date = input("Enter the date (YYYY-MM-DD): ")
        time = input("Enter the time (HH:MM): ")
        room = input("Enter the room number: ")    
        row = [task, date, time,room]
            
        sheet.append(row)
        workbook.save('tasks.xlsx')
            
main_program()
          
            
        
        